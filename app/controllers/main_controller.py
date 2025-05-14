from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from werkzeug.utils import secure_filename
import os
import logging
from app.models import User, UserActive, Active, Role, UserRole, Org, OrgUser, Log
from app.utils.auth import login_required, check_permission
from app.utils.pdf_parser import extract_org_info
from app import db
from flask_login import login_user, logout_user, current_user
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
from app.utils.pdf_utils import extract_bank_info_from_pdf
from app.models.org import Org
from app.models.org_config import OrgConfig
from app.utils.bank_utils import authenticate
from app.utils.email_utils import send_email 
# 配置日志
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

bp = Blueprint('main', __name__)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/')
def index():
    return render_template('index.html')

@bp.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        org_name = request.form.get('org_name')
        
        # 检查文件上传
        if 'bank_pdf' not in request.files:
            flash('请上传银行信息PDF文件', 'error')
            return redirect(url_for('main.register'))
        
        pdf_file = request.files['bank_pdf']
        if pdf_file.filename == '':
            flash('请选择PDF文件', 'error')
            return redirect(url_for('main.register'))
        
        # 保存PDF文件
        upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
        
        filename = secure_filename(pdf_file.filename)
        pdf_path = os.path.join(upload_folder, filename)
        pdf_file.save(pdf_path)
        
        # 从PDF中提取银行信息
         # 从PDF中提取银行信息和用户信息
        name, email, password, org_name, bank_name, bank_account, bank_password = extract_bank_info_from_pdf(pdf_path)
        
        if not all([name, email, password, org_name, bank_name, bank_account, bank_password]):
            flash('无法从PDF中提取完整的信息', 'error')
            return redirect(url_for('main.register'))
        try:
            org = Org(
                name=org_name,
                bank_account=bank_account,
                bank_name=bank_name,
                bank_password=bank_password,
            )
            success, message = authenticate(org)
            if not success:
                flash(f'认证失败: {message}', 'error')
                return redirect(url_for('main.register'))
            flash('认证成功', 'success')
            #return redirect(url_for('main.index'))

        except Exception as e:
            flash(f'注册过程出错: {str(e)}', 'error')
            return redirect(url_for('main.register'))

        try:
            # 创建组织
            org = Org(
                name=org_name,
                bank_name=bank_name,
                bank_account=bank_account,
                bank_password=bank_password,
            )
            db.session.add(org)
            db.session.commit()
            
            # 创建用户
            user = User(name=name, email=email)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            
            # 创建组织用户关系
            org_user = OrgUser(org_id=org.id, user_id=user.id)
            db.session.add(org_user)
            db.session.commit()

            eusers = UserRole.query.filter_by(role_id=2 or 3).all()
            euserIds = [euser.user_id for euser in eusers]
            eus=User.query.filter(User.id.in_(euserIds)).all()
            for euser in eus:
                send_email(euser.email, '审批通知', '有新的申请等待审核')
            send_email(user.email, '注册成功', '等待审核')
            # 登录用户
            flash('注册成功！', 'success')
            
            return redirect(url_for('main.index'))
            
        except Exception as e:
            db.session.rollback()
            flash('注册失败，请重试', 'error')
            return redirect(url_for('main.register'))

        flash('认证成功', 'success')
    return render_template('register.html')

@bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            if user.status == '0':
                flash('您的账号正在等待审批', 'warning')
                return render_template('login.html')
            elif user.status == '2':
                flash('您的账号已被拒绝', 'error')
                return render_template('login.html')
            
            # 记录登录日志
            try:
                log = Log(
                    log=f"用户 {user.name}({user.email}) 登录系统",
                    user_id=user.id
                )
                db.session.add(log)
                db.session.commit()
            except Exception as e:
                logger.error(f"记录登录日志失败: {str(e)}")
                db.session.rollback()
            
            login_user(user, remember=True)
            next_page = request.args.get('next')
            if not next_page or not next_page.startswith('/'):
                next_page = url_for('main.index')
            return redirect(next_page)
        else:
            flash('邮箱或密码错误', 'error')
    
    return render_template('login.html')

@bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('已退出登录')
    return redirect(url_for('main.index'))

@bp.route('/admin', methods=['GET', 'POST'])
@login_required
@check_permission(required_role='admin') 
def admin():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('请选择要上传的文件', 'error')
            return redirect(url_for('main.admin'))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('请选择要上传的文件', 'error')
            return redirect(url_for('main.admin'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('只允许上传Excel文件(.xlsx, .xls)', 'error')
            return redirect(url_for('main.admin'))
        
        try:
            # 读取Excel文件
            df = pd.read_excel(file)
            
            # 检查必要的列是否存在
            required_columns = ['邮箱', '角色']
            if not all(col in df.columns for col in required_columns):
                flash('Excel文件格式不正确，请确保包含：邮箱、角色列', 'error')
                return redirect(url_for('main.admin'))
            
            success_count = 0
            error_count = 0
            
            for _, row in df.iterrows():
                try:
                    email = str(row['邮箱']).strip()
                    role_value = str(row['角色']).strip().lower()
                    
                    if not email or not role_value:
                        logger.warning(f"Missing required fields for user {email}")
                        error_count += 1
                        continue
                    
                    if role_value not in ['e-admin', 'he-admin']:
                        logger.warning(f"Invalid role value for user {email}: {role_value}")
                        error_count += 1
                        continue
                    
                    # 检查用户是否存在
                    user = User.query.filter_by(email=email).first()
                    if not user:
                        # 创建新用户
                        user = User(
                            email=email,
                            name=email.split('@')[0],  # 临时使用邮箱前缀作为名称
                            status='1'  # 直接激活
                        )
                        # 设置默认密码为123456
                        user.set_password('123456')
                        db.session.add(user)
                        db.session.flush()
                        logger.info(f"Created new user: {email}")
                    
                    # 获取角色
                    role = Role.query.filter_by(value=role_value).first()
                    if not role:
                        role = Role(name='教务管理员' if role_value == 'e-admin' else '人事管理员', value=role_value)
                        db.session.add(role)
                        db.session.flush()
                    
                    # 检查是否已有角色关联
                    user_role = UserRole.query.filter_by(user_id=user.id, role_id=role.id).first()
                    if not user_role:
                        user_role = UserRole(user_id=user.id, role_id=role.id)
                        db.session.add(user_role)
                        success_count += 1
                    else:
                        logger.info(f"User {email} already has role {role_value}")
                
                except Exception as e:
                    logger.error(f"Error processing user {email}: {str(e)}")
                    error_count += 1
            
            db.session.commit()
            flash(f'导入完成：成功 {success_count} 条，失败 {error_count} 条', 'success' if success_count > 0 else 'warning')
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error importing users: {str(e)}")
            flash(f'导入失败：{str(e)}', 'error')
        
        return redirect(url_for('main.admin'))
    
    return render_template('admin.html')

@bp.route('/admin/template')
@login_required
@check_permission(required_role='admin')
def admin_template():
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "管理员导入模板"
        
        # 设置列名
        headers = ['邮箱', '角色']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # 添加示例数据
        sample_data = [
            ['example1@example.com', 'e-admin'],
            ['example2@example.com', 'he-admin']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # 保存临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # 发送文件
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name='admin_import_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"创建模板文件失败: {str(e)}")
        flash('创建模板失败，请重试', 'error')
        return redirect(url_for('main.admin'))
    finally:
        # 删除临时文件
        if 'temp_file' in locals():
            os.unlink(temp_file.name)

@bp.route('/orgs')
@login_required
def org_list():
    """显示所有组织列表"""
    orgs = Org.query.filter(Org.id != 0).all()
    return render_template('org_list.html', orgs=orgs)

@bp.route('/orgs/<int:org_id>/services')
@login_required
def org_services(org_id):
    """显示组织的服务列表"""
    org = Org.query.get_or_404(org_id)
    services = []
    
    # 获取组织配置的服务
    configs = OrgConfig.query.filter_by(org_id=org_id).all()
    
    # 服务类型映射
    service_types = {
        0: {'name': '论文服务', 'description': '提供论文相关服务'},
        1: {'name': '课程服务', 'description': '提供课程相关服务'},
        2: {'name': '学生认证', 'description': '提供学生认证服务'},
        3: {'name': '学生搜索', 'description': '提供学生搜索服务'}
    }
    
    # 构建服务列表
    for config in configs:
        service_type = service_types.get(config.feature_type, {'name': '未知服务', 'description': ''})
        services.append({
            'name': service_type['name'],
            'description': service_type['description'],
            'is_active': config.is_enabled,
            'feature_type': config.feature_type
        })
    
    return render_template('org_services.html', org=org, services=services) 