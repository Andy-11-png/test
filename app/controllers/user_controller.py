from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import login_required, current_user
from app import db
from app.models import User, Active, UserActive, UserAccount, Org, OrgUser, Role, UserRole, Log
import pandas as pd
import logging
from werkzeug.utils import secure_filename
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from app.utils.bank_utils import trans_money
from app.utils.email_utils import send_email
# 配置日志
logger = logging.getLogger(__name__)

bp = Blueprint('user', __name__)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/pending')
@login_required
def pending_users():
    logger.debug(f"Current user: {current_user.email}, is_e_admin: {current_user.is_he_admin}")
    logger.debug(f"User status: {current_user.status}")
    
    if not current_user.is_he_admin and not current_user.is_e_admin:
        logger.warning(f"User {current_user.email} attempted to access pending users without permission")
        flash('您没有权限访问此页面', 'error')
        return redirect(url_for('main.index'))
    
    # 添加更详细的查询日志
    if current_user.is_e_admin:
        pending_users = User.query.filter_by(status='0').all()
    elif current_user.is_he_admin:
        pending_users = User.query.filter(
            User.status == '1',
            User.he_admin_approver_id.is_(None),
            User.e_admin_approver_id.isnot(None)
        ).all()
    else:
        pending_users = []

    
    return render_template('user/pending.html', pending_users=pending_users)

@bp.route('/users/<int:user_id>/approve', methods=['POST'])
@login_required
def approve_user(user_id):
    """审批用户"""
    user = User.query.get_or_404(user_id)
    
    # 检查当前用户是否有审批权限
    if not (current_user.is_e_admin or current_user.is_he_admin):
        flash('您没有审批权限', 'error')
        return redirect(url_for('main.index'))
    
    try:
        if current_user.is_e_admin:
            # e-admin 审批
            if user.status == '0':  # 待审批
                user.status = '1'  # e-admin已审批
                user.e_admin_approver_id = current_user.id
                # 记录审批日志
                log = Log(
                    log=f"教务管理员 {current_user.name} 审批通过用户 {user.name}",
                    user_id=current_user.id
                )
                db.session.add(log)
                send_email(user.email, '审批通知', '您的申请已通过教务管理员审批，等待人事管理员审批')
                eusers = UserRole.query.filter_by(value='he-admin').all()   
                for euser in eusers:
                    send_email(euser.email, '审批通知', '有新的申请等待审核')
                flash('审批通过，等待人事管理员审批', 'success')
            else:
                flash('该用户已被审批过', 'warning')
        elif current_user.is_he_admin:
            # he-admin 审批
            if user.status == '1':  # e-admin已审批
                user.status = '2'  # he-admin已审批
                user.he_admin_approver_id = current_user.id
                # 记录审批日志
                log = Log(
                    log=f"人事管理员 {current_user.name} 审批通过用户 {user.name}",
                    user_id=current_user.id
                )
                db.session.add(log)
                flash('审批通过，用户已激活', 'success')
                send_email(user.email, '审批通知', '您的申请已通过人事管理员审批，用户已激活')
            elif user.status == '0':
                flash('请等待教务管理员审批', 'warning')
            else:
                flash('该用户已被审批过', 'warning')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'审批失败: {str(e)}', 'error')
    
    return redirect(url_for('user.pending_users'))

@bp.route('/users/<int:user_id>/reject', methods=['POST'])
@login_required
def reject_user(user_id):
    """拒绝用户"""
    user = User.query.get_or_404(user_id)
    
    # 检查当前用户是否有审批权限
    if not (current_user.is_e_admin or current_user.is_he_admin):
        flash('您没有审批权限', 'error')
        return redirect(url_for('main.index'))
    
    try:
        if current_user.is_e_admin:
            # e-admin 拒绝
            if user.status == '0':  # 待审批
                user.status = '3'  # 已拒绝
                user.e_admin_approver_id = current_user.id
                # 记录审批日志
                log = Log(
                    log=f"教务管理员 {current_user.name} 拒绝用户 {user.name}",
                    user_id=current_user.id
                )
                db.session.add(log)
                flash('已拒绝该用户', 'success')
            else:
                flash('该用户已被审批过', 'warning')
        elif current_user.is_he_admin:
            # he-admin 拒绝
            if user.status == '1':  # e-admin已审批
                user.status = '3'  # 已拒绝
                user.he_admin_approver_id = current_user.id
                # 记录审批日志
                log = Log(
                    log=f"人事管理员 {current_user.name} 拒绝用户 {user.name}",
                    user_id=current_user.id
                )
                db.session.add(log)
                flash('已拒绝该用户', 'success')
            elif user.status == '0':
                flash('请等待教务管理员审批', 'warning')
            else:
                flash('该用户已被审批过', 'warning')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'操作失败: {str(e)}', 'error')
    
    return redirect(url_for('user.pending_users'))

@bp.route('/org/<int:org_id>/users', methods=['GET', 'POST'])
@login_required
def org_users(org_id):
    org = Org.query.get_or_404(org_id)
    
    # 检查当前用户是否是组织的 convener
    org_user = OrgUser.query.filter_by(org_id=org_id, user_id=current_user.id).first()
    if not org_user:
        flash('您没有权限管理此组织', 'error')
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        
        user = User.query.get(user_id)
        if not user or not user.is_default:
            flash('无效的用户或用户角色不正确', 'error')
            return redirect(url_for('user.org_users', org_id=org_id))
        
        try:
            org_user = OrgUser(org_id=org_id, user_id=user_id)
            db.session.add(org_user)
            db.session.commit()
            flash('用户添加成功', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败: {str(e)}', 'error')
        
        return redirect(url_for('user.org_users', org_id=org_id))
    
    # 获取组织中的所有用户及其权限级别
    org_users_query = db.session.query(
        OrgUser, 
        User, 
        Active,
        UserAccount
    ).join(
        User, OrgUser.user_id == User.id
    ).outerjoin(
        UserActive, User.id == UserActive.user_id
    ).outerjoin(
        Active, UserActive.active_id == Active.id
    ).outerjoin(
        UserAccount, User.id == UserAccount.user_id
    ).filter(
        OrgUser.org_id == org_id
    ).all()

    # 处理查询结果
    org_users_data = []
    for org_user, user, active, account in org_users_query:
        org_users_data.append({
            'org_user': org_user,
            'user': user,
            'active_level': active.lv if active else None,
            'active_name': active.name if active else None,
            'quota': account.quota if account else 0
        })
    
    # 获取所有 default 角色的用户
    default_role = Role.query.filter_by(value='default').first()
    if not default_role:
        default_users = []
    else:
        default_users = User.query.join(UserRole).filter(
            UserRole.role_id == default_role.id,
            User.status == '1'
        ).all()
    
    return render_template('user/org_users.html', 
                         org=org, 
                         org_users=org_users_data,
                         default_users=default_users)

@bp.route('/org/<int:org_id>/users/<int:user_id>/update', methods=['POST'])
@login_required
def update_org_user(org_id, user_id):
    org = Org.query.get_or_404(org_id)
    
    # 检查当前用户是否是组织的 convener
    org_user = OrgUser.query.filter_by(org_id=org_id, user_id=current_user.id).first()
    if not org_user:
        flash('您没有权限管理此组织', 'error')
        return redirect(url_for('main.index'))
    
    active_level = int(request.form.get('active_level', 1))
    if not (1 <= active_level <= 3):
        flash('无效的权限级别', 'error')
        return redirect(url_for('user.org_users', org_id=org_id))
    
    try:
        # 获取目标用户
        target_user = User.query.get_or_404(user_id)
        
        # 获取新的权限级别对应的 active
        new_active = Active.query.filter_by(lv=active_level).first()
        if not new_active:
            flash('无效的权限级别', 'error')
            return redirect(url_for('user.org_users', org_id=org_id))
        
        # 删除用户现有的所有权限
        UserActive.query.filter_by(user_id=user_id).delete()
        
        # 添加新的权限
        user_active = UserActive(user_id=user_id, active_id=new_active.id)
        db.session.add(user_active)
        
        db.session.commit()
        flash('用户权限更新成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'更新失败: {str(e)}', 'error')
    
    return redirect(url_for('user.org_users', org_id=org_id))

@bp.route('/org/<int:org_id>/users/<int:user_id>/remove', methods=['POST'])
@login_required
def remove_org_user(org_id, user_id):
    org = Org.query.get_or_404(org_id)
    
    # 检查当前用户是否是组织的 convener
    org_user = OrgUser.query.filter_by(org_id=org_id, user_id=current_user.id).first()
    if not org_user:
        flash('您没有权限管理此组织', 'error')
        return redirect(url_for('main.index'))
    
    try:
        org_user = OrgUser.query.filter_by(org_id=org_id, user_id=user_id).first()
        if not org_user:
            flash('用户不在组织中', 'error')
            return redirect(url_for('user.org_users', org_id=org_id))
        
        db.session.delete(org_user)
        db.session.commit()
        flash('用户已从组织中移除', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'移除失败: {str(e)}', 'error')
    
    return redirect(url_for('user.org_users', org_id=org_id))

@bp.route('/my-org')
@login_required
def my_org():
    if not current_user.is_convener:
        flash('您没有权限访问此页面', 'error')
        return redirect(url_for('main.index'))
    
    # 获取当前用户所在的组织
    org_user = OrgUser.query.filter_by(user_id=current_user.id).first()
    if not org_user:
        flash('您不是任何组织的负责人', 'error')
        return redirect(url_for('main.index'))
    
    return redirect(url_for('user.org_users', org_id=org_user.org_id))

@bp.route('/org/<int:org_id>')
@login_required
def org_detail(org_id):
    org = Org.query.get_or_404(org_id)
    
    # 检查当前用户是否是组织的成员
    org_user = OrgUser.query.filter_by(org_id=org_id, user_id=current_user.id).first()
    if not org_user:
        flash('您没有权限访问此组织', 'error')
        return redirect(url_for('main.index'))
    
    # 获取组织中的所有用户及其权限级别
    org_users_query = db.session.query(
        OrgUser, 
        User, 
        Active,
        UserAccount
    ).join(
        User, OrgUser.user_id == User.id
    ).outerjoin(
        UserActive, User.id == UserActive.user_id
    ).outerjoin(
        Active, UserActive.active_id == Active.id
    ).outerjoin(
        UserAccount, User.id == UserAccount.user_id
    ).filter(
        OrgUser.org_id == org_id
    ).all()

    # 处理查询结果
    org_users_data = []
    for org_user, user, active, account in org_users_query:
        org_users_data.append({
            'org_user': org_user,
            'user': user,
            'active_level': active.lv if active else None,
            'active_name': active.name if active else None,
            'quota': account.quota if account else 0
        })
    
    return render_template('user/org_detail.html', 
                         org=org, 
                         org_users=org_users_data)

@bp.route('/org/<int:org_id>/users/import', methods=['POST'])
@login_required
def import_org_users(org_id):
    has_level1 = False
    money = 0 
    org = Org.query.get_or_404(org_id)
    
    # 检查当前用户是否是组织的 convener
    org_user = OrgUser.query.filter_by(org_id=org_id, user_id=current_user.id).first()
    if not org_user:
        flash('您没有权限管理此组织', 'error')
        return redirect(url_for('main.index'))
    
    if 'excel_file' not in request.files:
        flash('请选择要上传的文件', 'error')
        return redirect(url_for('user.org_users', org_id=org_id))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('请选择要上传的文件', 'error')
        return redirect(url_for('user.org_users', org_id=org_id))
    
    if not allowed_file(file.filename):
        flash('只允许上传Excel文件(.xlsx, .xls)', 'error')
        return redirect(url_for('user.org_users', org_id=org_id))
    
    try:
        # 读取Excel文件
        df = pd.read_excel(file)
        
        # 检查必要的列是否存在
        required_columns = ['邮箱', '姓名', '权限级别', '密码']
        if not all(col in df.columns for col in required_columns):
            flash('Excel文件格式不正确，请确保包含：邮箱、姓名、权限级别、密码列', 'error')
            return redirect(url_for('user.org_users', org_id=org_id))
        
        success_count = 0
        error_count = 0
        
        for _, row in df.iterrows():
            try:
                email = str(row['邮箱']).strip()
                name = str(row['姓名']).strip()
                active_level = int(row['权限级别'])
                password = str(row['密码']).strip()
                
                if not email or not name or not password:
                    logger.warning(f"Missing required fields for user {email}")
                    error_count += 1
                    continue
                
                if not (1 <= active_level <= 3):
                    logger.warning(f"Invalid active level for user {email}: {active_level}")
                    error_count += 1
                    continue
                
                # 检查用户是否已存在
                user = User.query.filter_by(email=email).first()
                if not user:
                    # 创建新用户
                    user = User(
                        email=email,
                        name=name,
                        status='1'  # 直接激活
                    )
                    user.set_password(password)
                    db.session.add(user)
                    db.session.flush()  # 获取user.id
                
                # 检查用户是否已在组织中
                existing_org_user = OrgUser.query.filter_by(org_id=org_id, user_id=user.id).first()
                if existing_org_user:
                    existing_org_user.active_level = active_level
                else:
                    # 创建组织用户关联
                    org_user = OrgUser(user_id=user.id, org_id=org_id)
                    db.session.add(org_user)
                
                # 处理active_level（权限级别）
                active = db.session.query(Active).filter_by(lv=active_level).first()
                if not active:
                    logger.error(f"Active level {active_level} not found for user {email}")
                    error_count += 1
                    continue
                
                # 检查是否已有关联，避免重复
                user_active = db.session.query(UserActive).filter_by(user_id=user.id, active_id=active.id).first()
                if not user_active:
                    user_active = UserActive(user_id=user.id, active_id=active.id)
                    db.session.add(user_active)
                
                # 设置用户角色
                role = Role.query.filter_by(value='default').first()
                if not role:
                    role = Role(name='普通用户', value='default')
                    db.session.add(role)
                    db.session.flush()
                
                # 检查是否已有角色关联
                user_role = UserRole.query.filter_by(user_id=user.id, role_id=role.id).first()
                if not user_role:
                    user_role = UserRole(user_id=user.id, role_id=role.id)
                    db.session.add(user_role)
                
                success_count += 1
                if active.lv == 1 and has_level1 == False:
                    has_level1 = True
                    money = 1000
                elif active.lv == 2 and has_level1 == False:
                    money+=1000
                elif active.lv == 3 and has_level1 == False:
                    money+=2000
            except Exception as e:
                logger.error(f"Error processing user {email}: {str(e)}")
                error_count += 1
        
        db.session.commit()
        flash(f'导入完成：成功 {success_count} 条，失败 {error_count} 条', 'success' if success_count > 0 else 'warning')
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error importing users: {str(e)}")
        flash(f'导入失败：{str(e)}', 'error')
    oorg = Org.query.filter_by(id=0).first()
    trans_money(org, oorg, money)
    return redirect(url_for('user.org_users', org_id=org_id))

@bp.route('/org/<int:org_id>/users/<int:user_id>/quota', methods=['POST'])
@login_required
def update_user_quota(org_id, user_id):
    org = Org.query.get_or_404(org_id)
    
    # 检查当前用户是否是组织的 convener
    org_user = OrgUser.query.filter_by(org_id=org_id, user_id=current_user.id).first()
    if not org_user:
        flash('您没有权限管理此组织', 'error')
        return redirect(url_for('main.index'))
    
    try:
        quota = int(request.form.get('quota', 0))
        if quota < 0:
            flash('额度不能为负数', 'error')
            return redirect(url_for('user.org_users', org_id=org_id))
        
        # 检查用户是否在组织中
        target_org_user = OrgUser.query.filter_by(org_id=org_id, user_id=user_id).first()
        if not target_org_user:
            flash('用户不在组织中', 'error')
            return redirect(url_for('user.org_users', org_id=org_id))
        
        # 获取或创建用户账户
        user_account = UserAccount.query.filter_by(user_id=user_id).first()
        if not user_account:
            user_account = UserAccount(user_id=user_id, quota=quota)
            db.session.add(user_account)
        else:
            user_account.quota = quota
        
        db.session.commit()
        flash('用户额度更新成功', 'success')
        
    except ValueError:
        flash('无效的额度值', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'更新失败: {str(e)}', 'error')
    
    return redirect(url_for('user.org_users', org_id=org_id))

@bp.route('/org/<int:org_id>/users/template')
@login_required
def download_org_user_template(org_id):
    """下载组织用户导入模板"""
    try:
        # 检查用户权限
        if not current_user.is_convener:
            flash('权限不足', 'error')
            return redirect(url_for('main.index'))
        
        # 检查用户是否属于该组织
        org_user = OrgUser.query.filter_by(user_id=current_user.id, org_id=org_id).first()
        if not org_user:
            flash('您不属于该组织', 'error')
            return redirect(url_for('main.index'))
        
        # 创建Excel文件
        output = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '用户导入模板'
        
        # 设置表头
        headers = ['邮箱', '姓名', '权限级别', '密码']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # 添加示例数据
        sample_data = [
            ['user1@example.com', '张三', '1', 'password123'],
            ['user2@example.com', '李四', '2', 'password456'],
            ['user3@example.com', '王五', '3', 'password789']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 30  # 邮箱
        sheet.column_dimensions['B'].width = 15  # 姓名
        sheet.column_dimensions['C'].width = 15  # 权限级别
        sheet.column_dimensions['D'].width = 20  # 密码
        
        
        # 保存文件
        workbook.save(output)
        output.seek(0)
        
        # 返回文件
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='用户导入模板.xlsx'
        )
    except Exception as e:
        logger.error(f"下载用户导入模板失败: {str(e)}")
        flash('下载模板失败', 'error')
        return redirect(url_for('main.index')) 

@bp.route('/user_org')
@login_required
def user_org():
    org_id = OrgUser.query.filter_by(user_id=current_user.id).first().org_id
    org = Org.query.filter_by(id=org_id).first()
    return jsonify({'name': org.name})